import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
import chardet
import csv
import sys
from typing import List, Set, Dict, Optional, Tuple
from urllib.parse import quote
import io
import base64

# Page configuration
st.set_page_config(
    page_title="BOM Tree Generator",
    page_icon="🌳",
    layout="wide",
    initial_sidebar_state="expanded"
)

def detect_file_encoding(file_path: str) -> str:
    """Detect file encoding using chardet"""
    with open(file_path, 'rb') as f:
        rawdata = f.read(10000)
    result = chardet.detect(rawdata)
    return result['encoding'] if result['confidence'] > 0.7 else 'utf-8'

def load_bom_files(bom_files: List[bytes], filenames: List[str]) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """Load multiple BOM files and return combined DF + individual DFs"""
    all_boms = []
    individual_boms = {}
    
    for file_data, filename in zip(bom_files, filenames):
        try:
            # Create a temporary file-like object
            file_obj = io.BytesIO(file_data)
            
            if filename.endswith('.txt') or filename.endswith('.csv'):
                # Try to detect encoding from the first part of the file
                file_obj.seek(0)
                rawdata = file_obj.read(10000)
                result = chardet.detect(rawdata)
                encoding = result['encoding'] if result['confidence'] > 0.7 else 'utf-8'
                
                file_obj.seek(0)
                df = pd.read_csv(
                    file_obj,
                    sep=None,
                    engine='python',
                    encoding=encoding,
                    on_bad_lines='skip',
                    dtype=str
                )
            else:
                file_obj.seek(0)
                df = pd.read_excel(file_obj, dtype=str)
            
            df['BOM_Source'] = filename
            individual_boms[filename] = df
            all_boms.append(df)
            st.success(f"✅ Loaded: {filename}")
        except Exception as e:
            st.error(f"❌ Error reading {filename}: {e}")
            continue
    
    if not all_boms:
        raise ValueError("No valid BOM files loaded")
    
    combined_bom = pd.concat(all_boms).drop_duplicates(
        subset=['Product no', 'Component no'],
        keep='first'
    ).fillna('')
    st.info(f"📊 Combined {len(all_boms)} BOMs → {len(combined_bom)} unique relationships")
    return combined_bom, individual_boms

def generate_drawing_url(part_number: str) -> str:
    """Generate M3 drawing URL based on part number"""
    base_url = "https://se-m3xiprod.systemair.se:9543/ca/client/index.html"
    query = f"/Drawing[@Drawing_number = \"{part_number}\"]"
    encoded_query = quote(query)
    return f"{base_url}?$pageNumber=1&%5BDrawing%20no%5D=&%27%22%5D=&$query={encoded_query}"

def build_tree(bom_df: pd.DataFrame, parent_part: str, level: int = 0, 
               seen_parts: Optional[Set[str]] = None) -> List[List[str]]:
    """Recursively build BOM tree with duplicate prevention"""
    seen_parts = seen_parts or set()
    
    if parent_part in seen_parts:
        return []
    seen_parts.add(parent_part)
    
    children = bom_df[bom_df['Product no'] == parent_part]
    rows = []

    for _, child in children.iterrows():
        component = str(child['Component no']).strip()
        if not component:
            continue
        
        if component in seen_parts:
            continue
            
        component_name = str(child.iloc[6]).strip() if len(child) > 6 else ""
        component_desc = str(child.iloc[7]).strip() if len(child) > 7 else ""
        bom_source = child.get('BOM_Source', 'Unknown')
        
        row = [''] * level + [component, component_name, component_desc, bom_source]
        rows.append(row)
        
        sub_rows = build_tree(bom_df, component, level + 1, seen_parts)
        rows.extend(sub_rows)

    return rows

def analyze_parts(bom_dfs: List[pd.DataFrame], parts: List[str], selected_sources: List[str]) -> Dict[str, List[List[str]]]:
    """Analyze list of parts across selected BOM sources"""
    all_trees = {}
    component_info = {}  # To store component info for parts without BOMs
    
    # First pass: find all components in the BOMs
    for bom_df in bom_dfs:
        for _, row in bom_df.iterrows():
            component = str(row['Component no']).strip()
            if component:
                component_info[component] = {
                    'name': str(row.iloc[6]).strip() if len(row) > 6 else "",
                    'description': str(row.iloc[7]).strip() if len(row) > 7 else "",
                    'source': row.get('BOM_Source', 'Unknown')
                }
    
    # Create progress bar and status container
    progress_bar = st.progress(0)
    status_container = st.empty()
    total_parts = len(parts)
    
    for i, part in enumerate(parts):
        part = str(part).strip()
        if not part:
            continue
        
        # Update status without creating new UI elements
        status_container.text(f"🔧 Processing: {part}")
        combined_tree = []
        seen_components = set()
        found_as_component = False
        
        # First try to find it as a parent (has its own BOM)
        for bom_df in bom_dfs:
            tree = build_tree(bom_df, part)
            if tree:
                # Merge trees while preventing duplicates
                for row in tree:
                    component = next((x for x in row if x.strip()), None)
                    if component and component not in seen_components:
                        combined_tree.append(row)
                        seen_components.add(component)
        
        if combined_tree:
            all_trees[part] = combined_tree
        else:
            # If not found as parent, check if it exists as a component
            if part in component_info:
                found_as_component = True
                # Create a single row with the component info
                info = component_info[part]
                all_trees[part] = [
                    [part, info['name'], info['description'], info['source']]
                ]
        
        progress_bar.progress((i + 1) / total_parts)
    
    # Clear the status container
    status_container.empty()
    progress_bar.empty()
    
    return all_trees

def create_excel_report(trees_dict: Dict[str, List[List[str]]], 
                       bom_df: pd.DataFrame, all_parts: List[str], selected_sources: List[str]) -> bytes:
    """Generate the final Excel report with formatting and auto-generated hyperlinks"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    
    hyperlink_style = NamedStyle(name='Hyperlink')
    hyperlink_style.font = Font(underline='single', color='0563C1')
    wb.add_named_style(hyperlink_style)
    
    no_bom_style = NamedStyle(name='NoBOM')
    no_bom_style.font = Font(color='FF0000')
    wb.add_named_style(no_bom_style)
    
    component_only_style = NamedStyle(name='ComponentOnly')
    component_only_style.font = Font(color='FFA500')  # Orange color
    wb.add_named_style(component_only_style)
    
    parts_with_boms = set(trees_dict.keys())
    parts_without_boms = set(all_parts) - parts_with_boms
    
    # Create worksheets for each part
    for part, rows in trees_dict.items():
        sheet_name = str(part).strip().replace(':', '_').replace('/', '_')[:31]
        ws_part = wb.create_sheet(sheet_name)
        
        # Get part info
        part_info = bom_df[bom_df['Product no'] == part]
        part_name = str(part_info.iloc[0, 2]) if not part_info.empty and len(part_info.iloc[0]) > 2 else ""
        
        # Add header
        if len(rows) == 1 and len(rows[0]) == 4:  # Component-only case
            ws_part.append(["Component Information"])
            ws_part.append(["Part Number:", part])
            ws_part.append(["Name:", rows[0][1]])
            ws_part.append(["Description:", rows[0][2]])
            ws_part.append(["Source BOM:", rows[0][3]])
            
            # Style the header
            for row in range(1, 6):
                ws_part[f"A{row}"].font = Font(bold=True)
                if row > 1:
                    ws_part[f"B{row}"].style = 'ComponentOnly'
            
            # Add hyperlink to component number
            if part.strip():
                cell = ws_part.cell(row=2, column=2)  # Part number cell
                cell.hyperlink = generate_drawing_url(part)
                cell.style = 'Hyperlink'
        else:
            # Regular BOM case
            ws_part.append(["Main Part:", part, part_name])
            ws_part.append([])
            ws_part.append(["Component", "Name", "Description", "Source BOM"])
            
            for col in ['A', 'B', 'C']:
                ws_part[f"{col}1"].font = Font(bold=True)
            
            for row in rows:
                indent_level = next((i for i, x in enumerate(row) if x != ''), 0)
                component_data = [x for x in row if x != ''][-4:]
                ws_part.append([''] * indent_level + component_data)
                
                component_col = indent_level + 1
                component_no = component_data[0].strip()
                if component_no:
                    cell = ws_part.cell(row=ws_part.max_row, column=component_col)
                    cell.hyperlink = generate_drawing_url(component_no)
                    cell.style = 'Hyperlink'
    
    # Create index sheet
    ws.append(["Part Number", "Part Name", "Status", "Source BOMs", "Searched In"])
    
    for part in all_parts:
        part_info = bom_df[bom_df['Product no'] == part]
        part_name = str(part_info.iloc[0, 2]) if not part_info.empty and len(part_info.iloc[0]) > 2 else ""
        
        if part in parts_with_boms:
            status = "BOM Found"
            sheet_name = str(part).strip().replace(':', '_').replace('/', '_')[:31]
            sources = ', '.join(sorted(set(part_info['BOM_Source']))) if 'BOM_Source' in part_info.columns else ""
            
            ws.append([part, part_name, status, sources, ', '.join(selected_sources)])
            ws[f"A{ws.max_row}"].hyperlink = f"#{sheet_name}!A1"
            ws[f"A{ws.max_row}"].style = 'Hyperlink'
        elif part in trees_dict and len(trees_dict[part]) == 1:  # Component-only
            status = "Component Only"
            sheet_name = str(part).strip().replace(':', '_').replace('/', '_')[:31]
            ws.append([part, part_name, status, trees_dict[part][0][3], ', '.join(selected_sources)])
            ws[f"A{ws.max_row}"].hyperlink = f"#{sheet_name}!A1"
            ws[f"A{ws.max_row}"].style = 'Hyperlink'
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f"{col}{ws.max_row}"].style = 'ComponentOnly'
        else:
            ws.append([part, part_name, "No BOM Found", "", ', '.join(selected_sources)])
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f"{col}{ws.max_row}"].style = 'NoBOM'
    
    # Set column widths
    for col, width in [('A', 20), ('B', 40), ('C', 15), ('D', 30), ('E', 30)]:
        ws.column_dimensions[col].width = width
    
    if len(wb.sheetnames) == 0:
        wb.create_sheet("Sheet1")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def load_predefined_boms() -> Dict[str, str]:
    """Return a dictionary of predefined BOM options with their GitHub URLs"""
    predefined_boms = {
        "1. Canada/467": "https://github.com/Harrybradrocco/Bomanalyze/raw/master/Canada.xlsx",
        "2. Spain/645": "https://github.com/Harrybradrocco/Bomanalyze/raw/master/Spain.xlsx", 
        "3. Denmark/770": "https://github.com/Harrybradrocco/Bomanalyze/raw/master/Denmark.xlsx"
    }
    return predefined_boms

def main():
    st.title("🌳 BOM Tree Generator")
    st.markdown("**SYSCAD TREE GENERATOR** - written by hbradroc@uwo.ca")
    st.markdown("---")
    
    # Sidebar for BOM selection
    st.sidebar.header("📁 Select BOM Files")
    
    # Predefined BOM options
    predefined_boms = load_predefined_boms()
    
    st.sidebar.subheader("Predefined BOM Files")
    
    # Show file status (always available from GitHub)
    for option in predefined_boms.keys():
        st.sidebar.success(f"✅ {option}")
    
    selected_predefined = st.sidebar.multiselect(
        "Choose predefined BOM files:",
        list(predefined_boms.keys()),
        default=list(predefined_boms.keys()),
        help="Select from the predefined BOM files"
    )
    
    st.sidebar.subheader("Custom BOM Files")
    uploaded_files = st.sidebar.file_uploader(
        "Upload additional BOM files (optional)",
        type=['xlsx', 'xls', 'csv', 'txt'],
        accept_multiple_files=True,
        help="Upload your own BOM files. Supported formats: Excel (.xlsx, .xls), CSV (.csv), Text (.txt)"
    )
    
    # Check if any BOM files are selected
    if not selected_predefined and not uploaded_files:
        st.info("👆 Please select predefined BOM files or upload custom files using the sidebar to get started.")
        st.markdown("""
        ### How to use this tool:
        1. **Select BOM Files**: Choose from predefined files or upload your own
        2. **Enter Part Numbers**: Add part numbers you want to analyze  
        3. **Select Sources**: Choose which BOM sources to search
        4. **Generate Report**: Download the Excel report with BOM trees
        
        ### Predefined BOM Files Available:
        - Canada/467
        - Spain/645  
        - Denmark/770
        
        ### Supported file formats:
        - Excel files (.xlsx, .xls)
        - CSV files (.csv)
        - Text files (.txt)
        """)
        return
    
    # Load BOM files
    try:
        all_file_data = []
        all_filenames = []
        
        # Load predefined files from GitHub
        for option in selected_predefined:
            github_url = predefined_boms[option]
            filename = github_url.split('/')[-1]  # Extract filename from URL
            
            try:
                import requests
                response = requests.get(github_url)
                response.raise_for_status()
                file_data = response.content
                all_file_data.append(file_data)
                all_filenames.append(filename)
                st.sidebar.success(f"✅ Loaded: {filename}")
            except Exception as e:
                st.sidebar.error(f"❌ Failed to load: {filename}")
                st.sidebar.error(f"Error: {str(e)}")
        
        # Load uploaded files
        if uploaded_files:
            for file in uploaded_files:
                all_file_data.append(file.getvalue())
                all_filenames.append(file.name)
        
        if not all_file_data:
            st.error("❌ No BOM files could be loaded. Please check file availability.")
            return
            
        combined_bom, individual_boms = load_bom_files(all_file_data, all_filenames)
        
        # Check required columns
        required_cols = ['Product no', 'Component no']
        missing = [col for col in required_cols if col not in combined_bom.columns]
        if missing:
            st.error(f"❌ Missing required columns: {', '.join(missing)}")
            st.info("Please ensure your BOM files contain 'Product no' and 'Component no' columns.")
            return
            
    except Exception as e:
        st.error(f"❌ Failed to load BOMs: {e}")
        return
    
    st.success("✅ BOM files loaded successfully!")
    
    # Main content area
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.header("🔍 Part Analysis")
        
        # Part number input
        st.subheader("Enter Part Numbers")
        input_method = st.radio(
            "Choose input method:",
            ["Manual Entry", "File Upload"],
            horizontal=True
        )
        
        parts = []
        
        if input_method == "Manual Entry":
            part_input = st.text_area(
                "Enter part numbers (one per line):",
                height=100,
                help="Enter one part number per line"
            )
            if part_input:
                parts = [p.strip() for p in part_input.split('\n') if p.strip()]
        
        else:  # File Upload
            part_file = st.file_uploader(
                "Upload file with part numbers",
                type=['xlsx', 'xls', 'csv', 'txt'],
                help="Upload a file containing part numbers. For Excel files, part numbers should be in column C."
            )
            
            if part_file:
                try:
                    if part_file.name.endswith(('.xlsx', '.xls')):
                        df = pd.read_excel(part_file, usecols="C", dtype=str)
                        parts = df.iloc[:, 0].dropna().unique().tolist()
                    else:  # CSV or TXT
                        # Try to detect encoding
                        file_content = part_file.getvalue()
                        result = chardet.detect(file_content[:10000])
                        encoding = result['encoding'] if result['confidence'] > 0.7 else 'utf-8'
                        
                        # Read the file
                        part_file.seek(0)
                        df = pd.read_csv(part_file, encoding=encoding, dtype=str)
                        # Assume part numbers are in the first column
                        parts = df.iloc[:, 0].dropna().unique().tolist()
                    
                    st.success(f"✅ Loaded {len(parts)} parts from file")
                except Exception as e:
                    st.error(f"❌ Error reading file: {e}")
        
        if parts:
            st.info(f"📋 Total parts to analyze: {len(parts)}")
            
            # Show first few parts
            with st.expander("Preview parts"):
                st.write(parts[:10])
                if len(parts) > 10:
                    st.write(f"... and {len(parts) - 10} more")
    
    with col2:
        st.header("⚙️ Settings")
        
        # BOM source selection
        st.subheader("Select BOM Sources")
        available_sources = list(individual_boms.keys())
        
        if len(available_sources) > 1:
            selected_sources = st.multiselect(
                "Choose BOM sources to search:",
                available_sources,
                default=available_sources,
                help="Select which BOM files to search for parts"
            )
        else:
            selected_sources = available_sources
            st.info(f"Using single BOM source: {available_sources[0]}")
        
        # Analysis button
        if st.button("🚀 Analyze Parts", type="primary", disabled=not parts):
            if not parts:
                st.warning("Please enter part numbers first.")
            elif not selected_sources:
                st.warning("Please select at least one BOM source.")
            else:
                # Perform analysis
                st.header("📊 Analysis Results")
                
                # Create containers for immediate display
                download_container = st.empty()
                summary_container = st.empty()
                
                # Show immediate feedback
                st.info("🔄 Analyzing parts... This may take a moment.")
                
                bom_dfs_to_search = [individual_boms[source] for source in selected_sources]
                all_trees = analyze_parts(bom_dfs_to_search, parts, selected_sources)
                
                if all_trees:
                    # Generate Excel report
                    excel_data = create_excel_report(all_trees, combined_bom, parts, selected_sources)
                    
                    # Show download button and summary immediately
                    with download_container:
                        st.download_button(
                            label="📥 Download Excel Report",
                            data=excel_data,
                            file_name=f"BOM_Analysis_Report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    with summary_container:
                        st.success("✅ Analysis complete!")
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Parts with BOMs", len(all_trees))
                        with col2:
                            st.metric("Parts without BOMs", len(parts) - len(all_trees))
                        with col3:
                            st.metric("Sources searched", len(selected_sources))
                else:
                    st.warning("❌ No valid BOM structures found for the provided parts.")

if __name__ == "__main__":
    main()
