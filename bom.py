import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
import chardet
import csv
import sys
from typing import List, Set, Dict, Optional, Tuple
from urllib.parse import quote

def install_package(package: str):
    """Install required package if not available"""
    try:
        __import__(package)
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

def load_predefined_boms() -> Dict[str, str]:
    """Return a dictionary of predefined BOM options with their file paths"""
    predefined_boms = {
        "1. Denmark/770": "Denmark.xlsx",
        "2. Spain/645": "Spain.xlsx",
        "3. Canada/467": "Canada.xlsx"
    }
    return predefined_boms

def select_bom_files() -> List[str]:
    """Let user select from predefined BOMs or enter custom paths"""
    predefined_boms = load_predefined_boms()
    print("\nAvailable predefined BOM files:")
    
    # Display predefined options
    for option, filename in predefined_boms.items():
        print(f"{option} - {filename}")
    
    print("4. Enter custom BOM file path")
    
    selected_paths = []
    while True:
        choices = input("\nSelect BOM files to use (comma-separated numbers 1-4): ").strip()
        
        try:
            selected_indices = [int(x.strip()) for x in choices.split(',')]
            if all(1 <= i <= 4 for i in selected_indices):
                # Handle predefined selections
                for i in selected_indices:
                    if i == 4:
                        # Custom file path entry
                        while True:
                            custom_path = input("\nEnter custom BOM file path (empty when done): ").strip()
                            if not custom_path:
                                break
                            if os.path.isfile(custom_path):
                                selected_paths.append(custom_path)
                                print(f"✅ Added: {os.path.basename(custom_path)}")
                            else:
                                print(f"❌ File not found: {custom_path}")
                    else:
                        # Predefined BOM
                        option = list(predefined_boms.keys())[i-1]
                        filename = predefined_boms[option]
                        if os.path.isfile(filename):
                            selected_paths.append(filename)
                            print(f"✅ Added predefined: {filename}")
                        else:
                            print(f"❌ Predefined file not found: {filename}")
                
                if not selected_paths:
                    print("❌ No BOM files selected")
                    continue
                
                return selected_paths
            else:
                print("❌ Invalid selection. Please enter numbers between 1-4.")
        except ValueError:
            print("❌ Invalid input. Please enter numbers separated by commas.")

def detect_file_encoding(file_path: str) -> str:
    """Detect file encoding using chardet"""
    with open(file_path, 'rb') as f:
        rawdata = f.read(10000)
    result = chardet.detect(rawdata)
    return result['encoding'] if result['confidence'] > 0.7 else 'utf-8'

def load_bom_files(bom_paths: List[str]) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """Load multiple BOM files and return combined DF + individual DFs"""
    all_boms = []
    individual_boms = {}
    for path in bom_paths:
        try:
            if path.endswith('.txt'):
                encoding = detect_file_encoding(path)
                print(f"🔍 Detected encoding: {encoding} for {os.path.basename(path)}")
                
                df = pd.read_csv(
                    path,
                    sep=None,
                    engine='python',
                    encoding=encoding,
                    on_bad_lines='skip',
                    dtype=str
                )
            else:
                df = pd.read_excel(path, dtype=str)
            
            df['BOM_Source'] = os.path.basename(path)
            individual_boms[os.path.basename(path)] = df
            all_boms.append(df)
            print(f"✅ Loaded: {os.path.basename(path)}")
        except Exception as e:
            print(f"❌ Error reading {path}: {e}")
            continue
    
    if not all_boms:
        raise ValueError("No valid BOM files loaded")
    
    combined_bom = pd.concat(all_boms).drop_duplicates(
        subset=['Product no', 'Component no'],
        keep='first'
    ).fillna('')
    print(f"📊 Combined {len(all_boms)} BOMs → {len(combined_bom)} unique relationships")
    return combined_bom, individual_boms

def generate_drawing_url(part_number: str) -> str:
    """Generate M3 drawing URL based on part number"""
    base_url = "https://se-m3xiprod.systemair.se:9543/ca/client/index.html"
    query = f"/Drawing[@Drawing_number = \"{part_number}\"]"
    encoded_query = quote(query)
    return f"{base_url}?$pageNumber=1&%5BDrawing%20no%5D=&%27%22%5D=&$query={encoded_query}"

def build_tree(bom_df: pd.DataFrame, parent_part: str, level: int = 0, 
               seen_parts: Optional[Set[str]] = None) -> List[List[str]]:
    """Recursively build BOM tree with duplicate prevention"""
    seen_parts = seen_parts or set()
    
    if parent_part in seen_parts:
        return []
    seen_parts.add(parent_part)
    
    children = bom_df[bom_df['Product no'] == parent_part]
    rows = []

    for _, child in children.iterrows():
        component = str(child['Component no']).strip()
        if not component:
            continue
        
        if component in seen_parts:
            continue
            
        component_name = str(child.iloc[6]).strip() if len(child) > 6 else ""
        component_desc = str(child.iloc[7]).strip() if len(child) > 7 else ""
        bom_source = child.get('BOM_Source', 'Unknown')
        
        row = [''] * level + [component, component_name, component_desc, bom_source]
        rows.append(row)
        
        sub_rows = build_tree(bom_df, component, level + 1, seen_parts)
        rows.extend(sub_rows)

    return rows

def analyze_parts(bom_dfs: List[pd.DataFrame], parts: List[str], selected_sources: List[str]) -> Dict[str, List[List[str]]]:
    """Analyze list of parts across selected BOM sources"""
    all_trees = {}
    component_info = {}  # To store component info for parts without BOMs
    
    # First pass: find all components in the BOMs
    for bom_df in bom_dfs:
        for _, row in bom_df.iterrows():
            component = str(row['Component no']).strip()
            if component:
                component_info[component] = {
                    'name': str(row.iloc[6]).strip() if len(row) > 6 else "",
                    'description': str(row.iloc[7]).strip() if len(row) > 7 else "",
                    'source': row.get('BOM_Source', 'Unknown')
                }
    
    for part in parts:
        part = str(part).strip()
        if not part:
            continue
        
        print(f"\n🔧 Processing: {part}")
        combined_tree = []
        seen_components = set()
        found_as_component = False
        
        # First try to find it as a parent (has its own BOM)
        for bom_df in bom_dfs:
            tree = build_tree(bom_df, part)
            if tree:
                # Merge trees while preventing duplicates
                for row in tree:
                    component = next((x for x in row if x.strip()), None)
                    if component and component not in seen_components:
                        combined_tree.append(row)
                        seen_components.add(component)
        
        if combined_tree:
            all_trees[part] = combined_tree
            print(f"✅ Found BOM structure with {len(combined_tree)} components across {len(bom_dfs)} sources")
        else:
            # If not found as parent, check if it exists as a component
            if part in component_info:
                found_as_component = True
                # Create a single row with the component info
                info = component_info[part]
                all_trees[part] = [
                    [part, info['name'], info['description'], info['source']]
                ]
                print(f"⚠️  Part found only as component - showing component information")
            else:
                print(f"❌ No BOM found for: {part} in selected sources")
    
    return all_trees

def save_to_excel(output_path: str, trees_dict: Dict[str, List[List[str]]], 
                 bom_df: pd.DataFrame, all_parts: List[str], selected_sources: List[str]):
    """Generate the final Excel report with formatting and auto-generated hyperlinks"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    
    hyperlink_style = NamedStyle(name='Hyperlink')
    hyperlink_style.font = Font(underline='single', color='0563C1')
    wb.add_named_style(hyperlink_style)
    
    no_bom_style = NamedStyle(name='NoBOM')
    no_bom_style.font = Font(color='FF0000')
    wb.add_named_style(no_bom_style)
    
    component_only_style = NamedStyle(name='ComponentOnly')
    component_only_style.font = Font(color='FFA500')  # Orange color
    wb.add_named_style(component_only_style)
    
    parts_with_boms = set(trees_dict.keys())
    parts_without_boms = set(all_parts) - parts_with_boms
    
    # Create worksheets for each part
    for part, rows in trees_dict.items():
        sheet_name = str(part).strip().replace(':', '_').replace('/', '_')[:31]
        ws_part = wb.create_sheet(sheet_name)
        
        # Get part info
        part_info = bom_df[bom_df['Product no'] == part]
        part_name = str(part_info.iloc[0, 2]) if not part_info.empty and len(part_info.iloc[0]) > 2 else ""
        
        # Add header
        if len(rows) == 1 and len(rows[0]) == 4:  # Component-only case
            ws_part.append(["Component Information"])
            ws_part.append(["Part Number:", part])
            ws_part.append(["Name:", rows[0][1]])
            ws_part.append(["Description:", rows[0][2]])
            ws_part.append(["Source BOM:", rows[0][3]])
            
            # Style the header
            for row in range(1, 6):
                ws_part[f"A{row}"].font = Font(bold=True)
                if row > 1:
                    ws_part[f"B{row}"].style = 'ComponentOnly'
            
            # Add hyperlink to component number
            if part.strip():
                cell = ws_part.cell(row=2, column=2)  # Part number cell
                cell.hyperlink = generate_drawing_url(part)
                cell.style = 'Hyperlink'
        else:
            # Regular BOM case
            ws_part.append(["Main Part:", part, part_name])
            ws_part.append([])
            ws_part.append(["Component", "Name", "Description", "Source BOM"])
            
            for col in ['A', 'B', 'C']:
                ws_part[f"{col}1"].font = Font(bold=True)
            
            for row in rows:
                indent_level = next((i for i, x in enumerate(row) if x != ''), 0)
                component_data = [x for x in row if x != ''][-4:]
                ws_part.append([''] * indent_level + component_data)
                
                component_col = indent_level + 1
                component_no = component_data[0].strip()
                if component_no:
                    cell = ws_part.cell(row=ws_part.max_row, column=component_col)
                    cell.hyperlink = generate_drawing_url(component_no)
                    cell.style = 'Hyperlink'
    
    # Create index sheet
    ws.append(["Part Number", "Part Name", "Status", "Source BOMs", "Searched In"])
    
    for part in all_parts:
        part_info = bom_df[bom_df['Product no'] == part]
        part_name = str(part_info.iloc[0, 2]) if not part_info.empty and len(part_info.iloc[0]) > 2 else ""
        
        if part in parts_with_boms:
            status = "BOM Found"
            sheet_name = str(part).strip().replace(':', '_').replace('/', '_')[:31]
            sources = ', '.join(sorted(set(part_info['BOM_Source']))) if 'BOM_Source' in part_info.columns else ""
            
            ws.append([part, part_name, status, sources, ', '.join(selected_sources)])
            ws[f"A{ws.max_row}"].hyperlink = f"#{sheet_name}!A1"
            ws[f"A{ws.max_row}"].style = 'Hyperlink'
        elif part in trees_dict and len(trees_dict[part]) == 1:  # Component-only
            status = "Component Only"
            sheet_name = str(part).strip().replace(':', '_').replace('/', '_')[:31]
            ws.append([part, part_name, status, trees_dict[part][0][3], ', '.join(selected_sources)])
            ws[f"A{ws.max_row}"].hyperlink = f"#{sheet_name}!A1"
            ws[f"A{ws.max_row}"].style = 'Hyperlink'
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f"{col}{ws.max_row}"].style = 'ComponentOnly'
        else:
            ws.append([part, part_name, "No BOM Found", "", ', '.join(selected_sources)])
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f"{col}{ws.max_row}"].style = 'NoBOM'
    
    # Set column widths
    for col, width in [('A', 20), ('B', 40), ('C', 15), ('D', 30), ('E', 30)]:
        ws.column_dimensions[col].width = width
    
    if len(wb.sheetnames) == 0:
        wb.create_sheet("Sheet1")
    
    wb.save(output_path)
    print(f"💾 Saved final report to {output_path}")

def read_text_file_with_retry(file_path: str, encoding: str, delimiter: str, 
                            part_col: int, max_retries: int = 3) -> Optional[List[str]]:
    """Read text file with retry logic for error handling"""
    parts = set()
    for attempt in range(max_retries):
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    try:
                        if len(row) > part_col:
                            part = str(row[part_col]).strip()
                            if part and part != 'nan':
                                parts.add(part)
                    except (IndexError, ValueError):
                        continue
            return list(parts)
        except UnicodeDecodeError as e:
            if attempt < max_retries - 1:
                encoding = 'utf-8'
                continue
            print(f"❌ Failed to read file after {max_retries} attempts: {e}")
            return None
        except Exception as e:
            print(f"❌ Unexpected error reading file: {e}")
            return None
    return None

def get_part_numbers_from_user() -> List[str]:
    """Continually prompt user for part numbers until they choose to stop"""
    all_parts = []
    while True:
        print("\nOptions:")
        print("1. Enter a part number")
        print("2. Import from file")
        print("3. Finish and process parts")
        
        choice = input("Select option (1-3): ").strip()
        
        if choice == '1':
            part = input("Enter part number: ").strip()
            if part:
                all_parts.append(part)
                print(f"Added part: {part}")
            else:
                print("❌ No part number entered")
        
        elif choice == '2':
            file_path = input("Enter file path (Make sure Parts are in Column C): ").strip()
            if not os.path.isfile(file_path):
                print("❌ File not found")
                continue
            
            try:
                if file_path.endswith('.txt'):
                    encoding = detect_file_encoding(file_path)
                    print(f"🔍 Detected encoding: {encoding}")
                    
                    with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                        sample = f.readline()
                        print(f"Sample first line: {sample[:200]}...")
                    
                    print("\nPlease specify:")
                    part_col = int(input("Part number column number (starting from 1): ")) - 1
                    delimiter = input("Delimiter (e.g., ',', ';', 'о', '|', 'Ï,'): ").strip()
                    
                    parts = read_text_file_with_retry(file_path, encoding, delimiter, part_col)
                    if parts:
                        all_parts.extend(parts)
                        print(f"✅ Added {len(parts)} parts from file")
                else:
                    df = pd.read_excel(file_path, usecols="C", dtype=str)
                    parts = df.iloc[:, 0].dropna().unique().tolist()
                    all_parts.extend(parts)
                    print(f"✅ Added {len(parts)} parts from file")
            except Exception as e:
                print(f"❌ Error reading file: {e}")
        
        elif choice == '3':
            if not all_parts:
                print("❌ No parts entered")
                continue
            break
        
        else:
            print("❌ Invalid option")
    
    # Remove duplicates while preserving order
    seen = set()
    unique_parts = [p for p in all_parts if not (p in seen or seen.add(p))]
    print(f"\nTotal unique parts to process: {len(unique_parts)}")
    return unique_parts

def select_bom_sources(individual_boms: Dict[str, pd.DataFrame]) -> List[str]:
    """Let user select which BOM sources to search"""
    print("\nAvailable BOM sources:")
    sources = list(individual_boms.keys())
    for i, source in enumerate(sources, 1):
        print(f"{i}. {source}")
    print(f"{len(sources)+1}. All sources")
    
    while True:
        selection = input("\nSelect sources to search (comma-separated numbers or 'all'): ").strip()
        if selection.lower() == 'all':
            return sources
        
        try:
            selected_indices = [int(x.strip()) for x in selection.split(',')]
            if all(1 <= i <= len(sources)+1 for i in selected_indices):
                if len(sources)+1 in selected_indices:
                    return sources
                return [sources[i-1] for i in selected_indices]
            print("❌ Invalid selection. Please try again.")
        except ValueError:
            print("❌ Invalid input. Please enter numbers separated by commas.")

def main():
    print("=== SYSCAD TREE GENERATOR ===")
    print("written by hbradroc@uwo.ca\n")
    
    install_package('chardet')
    install_package('openpyxl')
    install_package('pandas')
    
    # Load BOM files using new selection method
    bom_paths = select_bom_files()
    
    try:
        combined_bom, individual_boms = load_bom_files(bom_paths)
    except Exception as e:
        print(f"❌ Failed to load BOMs: {e}")
        return
    
    required_cols = ['Product no', 'Component no']
    missing = [col for col in required_cols if col not in combined_bom.columns]
    if missing:
        print(f"❌ Missing required columns: {', '.join(missing)}")
        return
    
    # Main interactive loop
    while True:
        print("\n=== Main Menu ===")
        print("1. Search parts")
        print("2. Exit")
        
        choice = input("Select option (1-2): ").strip()
        
        if choice == '1':
            # Let user select which BOMs to search
            selected_sources = select_bom_sources(individual_boms)
            bom_dfs_to_search = [individual_boms[source] for source in selected_sources]
            
            parts = get_part_numbers_from_user()
            if not parts:
                continue
            
            all_trees = analyze_parts(bom_dfs_to_search, parts, selected_sources)
            if not all_trees:
                print("❌ No valid BOM structures generated")
                continue
            
            output_file = f"BOM_Analysis_Report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_to_excel(output_file, all_trees, combined_bom, parts, selected_sources)
            
            print("\n=== Analysis Complete ===")
            print(f"Sources searched: {', '.join(selected_sources)}")
            print(f"Total parts processed: {len(parts)}")
            print(f"Parts with BOMs: {len(all_trees)}")
            print(f"Parts without BOMs: {len(parts) - len(all_trees)}")
            print(f"Report generated: {output_file}")
        
        elif choice == '2':
            print("Exiting program...")
            break
        
        else:
            print("❌ Invalid option")

if __name__ == "__main__":
    main()